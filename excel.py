import openpyxl
from openpyxl import Workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.mime.text import MIMEText
import os
from datetime import datetime

input_wb = openpyxl.load_workbook('output.xlsx')
input_ws = input_wb.active

headers = [cell.value for cell in input_ws[1]]
sno_index = headers.index('SNo')
email_index = headers.index('Email')

smtp_server = 'smtp.gmail.com'
smtp_port = 587
smtp_user = 'neetasingh18011988@gmail.com'
smtp_password = 'Oshiraghav@31'  

today_date = datetime.now().strftime('%Y-%m-%d')
if not os.path.exists(today_date):
    os.makedirs(today_date)

def send_email(to_email, subject, body, attachment_path):
    msg = MIMEMultipart()
    msg['From'] = smtp_user
    msg['To'] = to_email
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'plain'))

    attachment = MIMEBase('application', 'octet-stream')
    with open(attachment_path, 'rb') as attachment_file:
        attachment.set_payload(attachment_file.read())
    encoders.encode_base64(attachment)
    attachment.add_header('Content-Disposition', f'attachment; filename={attachment_path}')
    msg.attach(attachment)

    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.set_debuglevel(1)  # Enable debug output
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.sendmail(smtp_user, to_email, msg.as_string())
        print(f'Email sent to {to_email}')
    except Exception as e:
        print(f'Failed to send email to {to_email}: {e}')

for row in input_ws.iter_rows(min_row=2, values_only=True):
    output_wb = Workbook()
    output_ws = output_wb.active

    for col_index, column_name in enumerate(headers, start=1):
        output_ws.cell(row=1, column=col_index, value=column_name)

    for col_index, cell_value in enumerate(row, start=1):
        output_ws.cell(row=2, column=col_index, value=cell_value)

    sno_value = row[sno_index]
    output_filename = os.path.join(today_date, f'{sno_value}.xlsx')
    output_wb.save(output_filename)
    print(f'Saved {output_filename}')

    email_address = row[email_index]
    email_subject = f"Your Data - SNo {sno_value}" 
    email_body = f"Please find attached the data for SNo {sno_value}."
    send_email(email_address, email_subject, email_body, output_filename)
    print(f'Sent email to {email_address} with attachment {output_filename}')
